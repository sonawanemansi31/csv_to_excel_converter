import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# -----------------------------
# Logging Setup
# -----------------------------
import tempfile

# -----------------------------
# Safe Logging Setup
# -----------------------------
LOG_DIR = os.path.join(tempfile.gettempdir(), "csv_to_excel_logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, "converter.log")

try:
    logging.basicConfig(
        filename=LOG_FILE,
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s"
    )
except PermissionError:
    # Fallback to console logging if file access fails
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s"
    )
    print("Warning: Could not create log file. Using console logging only.")


class CSVToExcelProApp:
    def __init__(self, root):
        self.root = root
        self.root.title("CSV to Excel Converter PRO")
        self.root.geometry("1100x700")
        self.root.configure(bg="#1e1e2f")
        self.root.minsize(1000, 650)

        self.selected_files = []
        self.output_folder = tk.StringVar()
        self.rename_columns = tk.StringVar()
        self.preview_df = None

        self.setup_styles()
        self.create_widgets()

    # -----------------------------
    # Styling
    # -----------------------------
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("TFrame", background="#1e1e2f")
        style.configure("TLabel", background="#1e1e2f", foreground="white", font=("Segoe UI", 10))
        style.configure("Header.TLabel", font=("Segoe UI", 20, "bold"), foreground="#00d4ff")
        style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=8)
        style.configure("Treeview",
                        background="#2d2d44",
                        foreground="white",
                        fieldbackground="#2d2d44",
                        rowheight=28)
        style.configure("Treeview.Heading",
                        font=("Segoe UI", 10, "bold"),
                        background="#00a8cc",
                        foreground="white")
        style.map("Treeview", background=[("selected", "#4CAF50")])

    # -----------------------------
    # UI Layout
    # -----------------------------
    def create_widgets(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Title
        title = ttk.Label(main_frame, text="CSV to Excel Converter PRO", style="Header.TLabel")
        title.pack(pady=10)

        # Buttons frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=10)

        tk.Button(btn_frame, text="Select CSV Files", command=self.select_files,
                  bg="#4CAF50", fg="white", font=("Segoe UI", 10, "bold"), width=18).pack(side="left", padx=5)

        tk.Button(btn_frame, text="Select Output Folder", command=self.select_output_folder,
                  bg="#2196F3", fg="white", font=("Segoe UI", 10, "bold"), width=18).pack(side="left", padx=5)

        tk.Button(btn_frame, text="Preview First File", command=self.preview_file,
                  bg="#9C27B0", fg="white", font=("Segoe UI", 10, "bold"), width=18).pack(side="left", padx=5)

        tk.Button(btn_frame, text="Convert All Files", command=self.convert_all_files,
                  bg="#FF9800", fg="white", font=("Segoe UI", 10, "bold"), width=18).pack(side="left", padx=5)

        tk.Button(btn_frame, text="Open Log File", command=self.open_log_file,
                  bg="#607D8B", fg="white", font=("Segoe UI", 10, "bold"), width=15).pack(side="left", padx=5)

        # File info
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill="x", pady=10)

        self.files_label = ttk.Label(info_frame, text="Selected Files: None")
        self.files_label.pack(anchor="w", pady=3)

        self.output_label = ttk.Label(info_frame, text="Output Folder: None")
        self.output_label.pack(anchor="w", pady=3)

        # Rename columns
        rename_frame = ttk.Frame(main_frame)
        rename_frame.pack(fill="x", pady=10)

        ttk.Label(rename_frame, text='Rename Columns (optional): old1:new1,old2:new2').pack(anchor="w")
        tk.Entry(rename_frame, textvariable=self.rename_columns, width=80,
                 font=("Segoe UI", 10), bg="#2d2d44", fg="white", insertbackground="white").pack(fill="x", pady=5)

        # Preview Table
        preview_frame = ttk.Frame(main_frame)
        preview_frame.pack(fill="both", expand=True, pady=10)

        ttk.Label(preview_frame, text="Data Preview (First 10 Rows)").pack(anchor="w")

        table_frame = ttk.Frame(preview_frame)
        table_frame.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(table_frame, show="headings")
        self.tree.pack(side="left", fill="both", expand=True)

        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scrollbar_y.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar_y.set)

        scrollbar_x = ttk.Scrollbar(preview_frame, orient="horizontal", command=self.tree.xview)
        scrollbar_x.pack(fill="x")
        self.tree.configure(xscrollcommand=scrollbar_x.set)

        # Status box
        ttk.Label(main_frame, text="Status / Logs").pack(anchor="w", pady=(10, 0))
        self.status_box = tk.Text(main_frame, height=8, bg="#111827", fg="#00ff99",
                                  font=("Consolas", 10), insertbackground="white")
        self.status_box.pack(fill="x", pady=5)

    # -----------------------------
    # Logging to GUI + file
    # -----------------------------
    def log(self, message, level="info"):
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted = f"[{timestamp}] {message}\n"
        self.status_box.insert(tk.END, formatted)
        self.status_box.see(tk.END)

        if level == "info":
            logging.info(message)
        elif level == "error":
            logging.error(message)
        elif level == "warning":
            logging.warning(message)

    # -----------------------------
    # File Selection
    # -----------------------------
    def select_files(self):
        files = filedialog.askopenfilenames(
            title="Select CSV Files",
            filetypes=[("CSV Files", "*.csv")]
        )
        if files:
            self.selected_files = list(files)
            self.files_label.config(text=f"Selected Files: {len(self.selected_files)} file(s)")
            self.log(f"Selected {len(self.selected_files)} CSV file(s).")

    def select_output_folder(self):
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            self.output_folder.set(folder)
            self.output_label.config(text=f"Output Folder: {folder}")
            self.log(f"Selected output folder: {folder}")

    # -----------------------------
    # Column Rename Parsing
    # -----------------------------
    def parse_rename_columns(self, rename_str):
        rename_dict = {}
        if rename_str.strip():
            pairs = rename_str.split(",")
            for pair in pairs:
                if ":" in pair:
                    old, new = pair.split(":", 1)
                    rename_dict[old.strip()] = new.strip()
        return rename_dict

    # -----------------------------
    # Data Cleaning
    # -----------------------------
    def clean_dataframe(self, df):
        df.columns = [str(col).strip() for col in df.columns]

        for col in df.select_dtypes(include=["object"]).columns:
            df[col] = df[col].astype(str).str.strip()

        df.replace("nan", pd.NA, inplace=True)

        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                df[col] = df[col].fillna(0)
            else:
                df[col] = df[col].fillna("Unknown")

        return df

    # -----------------------------
    # Date Parsing
    # -----------------------------
    def parse_dates(self, df):
        for col in df.columns:
            if "date" in col.lower() or "time" in col.lower():
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                    self.log(f"Parsed date column: {col}")
                except Exception as e:
                    self.log(f"Could not parse date column '{col}': {e}", "warning")
        return df

    # -----------------------------
    # Preview First File
    # -----------------------------
    def preview_file(self):
        if not self.selected_files:
            messagebox.showwarning("Warning", "Please select at least one CSV file first.")
            return

        try:
            file_path = self.selected_files[0]
            self.preview_df = pd.read_csv(file_path)
            self.preview_df = self.clean_dataframe(self.preview_df)
            self.preview_df = self.parse_dates(self.preview_df)

            self.show_dataframe_in_tree(self.preview_df.head(10))
            self.log(f"Preview loaded for: {os.path.basename(file_path)}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to preview file:\n{e}")
            self.log(f"Preview error: {e}", "error")

    def show_dataframe_in_tree(self, df):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = list(df.columns)

        for col in df.columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150, anchor="center")

        for _, row in df.iterrows():
            self.tree.insert("", "end", values=list(row))

    # -----------------------------
    # Excel Formatting
    # -----------------------------
    def format_excel(self, file_path):
        wb = load_workbook(file_path)
        ws = wb.active

        # Header formatting
        header_fill = PatternFill(start_color="00A8CC", end_color="00A8CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(file_path)

    # -----------------------------
    # Convert All Files
    # -----------------------------
    def convert_all_files(self):
        if not self.selected_files:
            messagebox.showerror("Error", "Please select CSV files.")
            return

        if not self.output_folder.get():
            messagebox.showerror("Error", "Please select an output folder.")
            return

        rename_dict = self.parse_rename_columns(self.rename_columns.get())
        success_count = 0

        for file_path in self.selected_files:
            try:
                self.log(f"Processing: {os.path.basename(file_path)}")

                df = pd.read_csv(file_path)
                df = self.clean_dataframe(df)
                df = self.parse_dates(df)

                if rename_dict:
                    df.rename(columns=rename_dict, inplace=True)
                    self.log(f"Applied column renames to {os.path.basename(file_path)}")

                base_name = os.path.splitext(os.path.basename(file_path))[0]
                output_path = os.path.join(self.output_folder.get(), f"{base_name}.xlsx")

                df.to_excel(output_path, index=False, engine="openpyxl")
                self.format_excel(output_path)

                self.log(f"Saved: {output_path}")
                success_count += 1

            except pd.errors.EmptyDataError:
                self.log(f"Skipped empty file: {os.path.basename(file_path)}", "warning")

            except pd.errors.ParserError:
                self.log(f"Invalid CSV format: {os.path.basename(file_path)}", "error")

            except PermissionError:
                self.log(f"Permission denied (file may be open): {os.path.basename(file_path)}", "error")

            except Exception as e:
                self.log(f"Unexpected error in {os.path.basename(file_path)}: {e}", "error")

        messagebox.showinfo("Completed", f"Conversion finished!\nSuccessfully converted: {success_count}/{len(self.selected_files)} files")
        self.log(f"Conversion completed: {success_count}/{len(self.selected_files)} files converted.")

    # -----------------------------
    # Open Log File
    # -----------------------------
    def open_log_file(self):
        try:
           if os.path.exists(LOG_FILE):
            os.startfile(LOG_FILE)  # Windows only
           else:
            messagebox.showwarning("Warning", "Log file not found yet.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open log file:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = CSVToExcelProApp(root)
    root.mainloop()